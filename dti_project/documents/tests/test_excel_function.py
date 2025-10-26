# ===== TEST FUNCTION =====
def test_export_format():
    """
    Standalone test function to generate a sample Excel file
    Run this independently to see the output format
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import datetime
    
    # Create sample data
    class MockItem:
        def __init__(self, date, number, code, payer, particulars, amount):
            self.date = date
            self.number = number
            self.responsibility_code = code
            self.payer = payer
            self.particulars = particulars
            self.amount = amount
    
    class MockReport:
        def __init__(self):
            self.pk = 1
            self.dti_office = "DTI Albay Provincial Office"
            self.report_no = "25-001"
            self.report_collection_date = datetime.date(2025, 1, 2)
            self.certification = "I hereby certify that the above is a true and correct report of collections made by me during the period covered by which Official Receipt Nos. 8722481 to 8722481 and BN-25-0140-0061-0009411-0009444 issued by me in the amounts shown therein. I also certify that I have not received money from whatever source, saving receipt, issuing receipt. The necessary Official Receipts in acknowledgement thereof. Collections received by me is/are/was/were deposited in/partly deposited as/only kept if/ever and shall be deposited their proper collection report numbers. I certify further that the balance shown above agrees with the balance appearing in my Cash Receipts Record."
            self.name_of_collection_officer = "Ma. Lourdes A. Pasobillo"
            self.official_designation = "Special Collecting Officer"
            
            # Sample items
            self.items = [
                MockItem(
                    datetime.date(2025, 1, 2),
                    "BN-25-0140-0061-0009411-0009444",
                    "27C-11-09-21",
                    "BERITA, MARJAM JANE MA, LADINGAG",
                    "CANCELLATION OF BUSINESS NAME REGISTRATION",
                    30.00
                ),
                MockItem(
                    datetime.date(2025, 1, 2),
                    "BN-25-0140-0061-0009411-0009444",
                    "",
                    "CAPACITE, FE ABNE BUAC",
                    "CANCELLATION OF BUSINESS NAME",
                    30.00
                ),
                MockItem(
                    datetime.date(2025, 1, 3),
                    "BN-25-0140-0061-0009411-0009448",
                    "",
                    "BENASAYAG, IMELDA SIRETE",
                    "RENEWAL OF BUSINESS NAME",
                    1000.00
                ),
                MockItem(
                    datetime.date(2025, 1, 3),
                    "BN-25-0140-0061-0009411-0009449",
                    "",
                    "ATIENZA NO, JUVITH MUNICHO",
                    "CANCELLATION OF BUSINESS",
                    30.00
                ),
                MockItem(
                    datetime.date(2025, 1, 5),
                    "8722481",
                    "",
                    "R-3 CYCLE STORE SPARE PARTS & ACCESSORIES",
                    "RENEWAL OF ACCREDITATION",
                    430.00
                ),
            ]
    
    report = MockReport()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    sheet_name = f"Report {report.report_no}"
    ws = wb.create_sheet(title=sheet_name)
    
    current_row = 1
    
    # --- HEADER SECTION ---
    ws.merge_cells(f'A{current_row}:H{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = "REPORT OF COLLECTIONS AND DEPOSITS"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_alignment
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:H{current_row}')
    office_cell = ws[f'A{current_row}']
    office_cell.value = report.dti_office
    office_cell.font = Font(bold=True)
    office_cell.alignment = center_alignment
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:H{current_row}')
    date_cell = ws[f'A{current_row}']
    date_cell.value = report.report_collection_date
    date_cell.alignment = center_alignment
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:H{current_row}')
    report_no_cell = ws[f'A{current_row}']
    report_no_cell.value = f"Report No. {report.report_no}"
    report_no_cell.alignment = center_alignment
    current_row += 1
    
    current_row += 1
    
    # --- TABLE HEADER ---
    header_row = current_row
    
    ws.merge_cells(f'A{header_row}:B{header_row}')
    ws[f'A{header_row}'] = "Official Receipt"
    ws[f'C{header_row}'] = "Responsibility"
    ws[f'D{header_row}'] = ""
    ws[f'E{header_row}'] = ""
    ws[f'F{header_row}'] = "Particulars"
    ws[f'G{header_row}'] = ""
    ws[f'H{header_row}'] = "Amount"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws[f'{col}{header_row}']
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border
    
    current_row += 1
    
    ws[f'A{current_row}'] = "Date"
    ws[f'B{current_row}'] = "Number"
    ws[f'C{current_row}'] = "Code"
    ws[f'D{current_row}'] = ""
    ws[f'E{current_row}'] = "Payer"
    ws[f'F{current_row}'] = ""
    ws[f'G{current_row}'] = ""
    ws[f'H{current_row}'] = ""
    
    for col in ['A', 'B', 'C', 'E', 'H']:
        cell = ws[f'{col}{current_row}']
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border
    
    current_row += 1
    
    # --- DATA ROWS ---
    for item in report.items:
        ws[f'A{current_row}'] = item.date
        ws[f'B{current_row}'] = item.number
        ws[f'C{current_row}'] = item.responsibility_code
        ws[f'D{current_row}'] = ""
        ws[f'E{current_row}'] = item.payer
        ws[f'F{current_row}'] = item.particulars
        ws[f'G{current_row}'] = ""
        ws[f'H{current_row}'] = item.amount
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            cell = ws[f'{col}{current_row}']
            cell.border = thin_border
            if col in ['A', 'B', 'C']:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 'H':
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00'
        
        current_row += 1
    
    # --- SUMMARY SECTION ---
    current_row += 2
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "Summary:"
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = "Undeposited Collections per last Report"
    ws[f'F{current_row}'] = "P"
    ws[f'H{current_row}'] = 1210.00
    ws[f'H{current_row}'].number_format = '#,##0.00'
    current_row += 2
    
    or_numbers = [item.number for item in report.items if item.number]
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = f"Collections per OR Nos. 8722481 to 8722481"
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = "and BN-25-0140-0061-0009411-0009444 to"
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = "BN-25-0140-0061-0009415-0009448"
    ws[f'H{current_row}'] = 1520.00
    ws[f'H{current_row}'].number_format = '#,##0.00'
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = "Total"
    ws[f'A{current_row}'].font = bold_font
    ws[f'H{current_row}'] = 2730.00
    ws[f'H{current_row}'].number_format = '#,##0.00'
    ws[f'H{current_row}'].font = bold_font
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = "Deposits"
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    # Date deposits row - don't merge, write to individual cells
    ws[f'A{current_row}'] = "Date"
    ws[f'B{current_row}'] = "2-Jan-25"
    ws[f'H{current_row}'] = 1000.00
    ws[f'H{current_row}'].number_format = '#,##0.00'
    current_row += 1
    
    ws[f'H{current_row}'] = 120.00
    ws[f'H{current_row}'].number_format = '#,##0.00'
    current_row += 1
    
    ws[f'H{current_row}'] = 90.00
    ws[f'H{current_row}'].number_format = '#,##0.00'
    current_row += 1
    
    ws[f'H{current_row}'] = 1400.00
    ws[f'H{current_row}'].number_format = '#,##0.00'
    current_row += 2
    
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = "Undeposited Collections, this Report"
    ws[f'F{current_row}'] = "P"
    ws[f'H{current_row}'] = 2510.00
    ws[f'H{current_row}'].number_format = '#,##0.00'
    ws[f'H{current_row}'].font = bold_font
    current_row += 1
    
    ws[f'H{current_row}'] = 180.00
    ws[f'H{current_row}'].number_format = '#,##0.00'
    current_row += 2
    
    # --- CERTIFICATION SECTION ---
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cert_cell = ws[f'A{current_row}']
    cert_cell.value = "CERTIFICATION"
    cert_cell.font = Font(bold=True, size=11)
    cert_cell.alignment = center_alignment
    current_row += 2
    
    ws.merge_cells(f'A{current_row}:H{current_row + 3}')
    cert_text_cell = ws[f'A{current_row}']
    cert_text_cell.value = report.certification
    cert_text_cell.alignment = Alignment(wrap_text=True, vertical="top")
    current_row += 5
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    name_cell = ws[f'A{current_row}']
    name_cell.value = report.name_of_collection_officer
    name_cell.font = bold_font
    name_cell.alignment = center_alignment
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "Name and Signature of Collecting Officer"
    ws[f'A{current_row}'].alignment = center_alignment
    current_row += 2
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    designation_cell = ws[f'A{current_row}']
    designation_cell.value = report.official_designation
    designation_cell.font = bold_font
    designation_cell.alignment = center_alignment
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "Official Designation"
    ws[f'A{current_row}'].alignment = center_alignment
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    date_cert_cell = ws[f'A{current_row}']
    date_cert_cell.value = report.report_collection_date
    date_cert_cell.alignment = center_alignment
    
    # --- SET COLUMN WIDTHS ---
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 5
    ws.column_dimensions['H'].width = 15
    
    # Save to Downloads folder
    import os
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
    output_path = os.path.join(downloads_path, "Sample_Collection_Report.xlsx")
    wb.save(output_path)
    print(f"✓ Test file generated: {output_path}")
    print(f"✓ Report: {report.report_no}")
    print(f"✓ Office: {report.dti_office}")
    print(f"✓ Date: {report.report_collection_date}")
    print(f"✓ Items: {len(report.items)}")
    print(f"✓ Officer: {report.name_of_collection_officer}")
    print(f"✓ Designation: {report.official_designation}")
    return output_path


# Run test if executed directly
if __name__ == "__main__":
    test_export_format()