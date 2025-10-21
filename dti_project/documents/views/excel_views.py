import datetime
from io import BytesIO
from itertools import chain
import json
import re
import threading
import time
from urllib.parse import quote
import uuid
from django.urls import reverse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, HttpResponseRedirect, StreamingHttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from ..models.collection_models import CollectionReport, CollectionReportItem
from ..mappings import EXPORT_MODEL_MAP, UPLOAD_MODEL_MAP
from ..utils.excel_view_helpers import get_model_from_sheet, get_user_from_full_name, normalize_header, normalize_sheet_name, shorten_name, to_title
from django.contrib.auth.mixins import LoginRequiredMixin
from ..mixins.permissions_mixins import UserRoleMixin
from django.shortcuts import redirect, render
from django.contrib import messages
from django.db import IntegrityError, transaction
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from django.utils.dateparse import parse_date
from django.http import JsonResponse
from django.core.cache import cache

import os
from ..models import (
    ChecklistEvaluationSheet,
    InspectionValidationReport,
    OrderOfPayment,
    PersonalDataSheet,
    SalesPromotionPermitApplication,
    ServiceRepairAccreditationApplication,
)


class ExportDocumentsExcelView(LoginRequiredMixin, View):
    """
    Export selected documents (via checkboxes) or all filtered documents to Excel.
    Preserves grouped formatting, headers, styling, and filename date ranges.
    """

    def post(self, request, *args, **kwargs):
        user = request.user
        selected_docs = request.POST.getlist("documents")

        # --- Collect documents by model ---
        grouped = {}
        if selected_docs:
            # Use checkbox selection
            grouped_ids = {}
            for doc in selected_docs:
                try:
                    model_name, pk = doc.split(":")
                    grouped_ids.setdefault(model_name.lower(), []).append(pk)
                except ValueError:
                    continue

            for model_name, ids in grouped_ids.items():
                model = EXPORT_MODEL_MAP.get(model_name)
                if not model:
                    continue
                qs = UserRoleMixin.get_queryset_or_all(model, user).filter(pk__in=ids)
                if qs.exists():
                    grouped[model._meta.verbose_name_plural.title()] = list(qs)
        else:
            # If no checkboxes, export all (apply UserRoleMixin)
            for model in EXPORT_MODEL_MAP.values():
                qs = UserRoleMixin.get_queryset_or_all(model, user)
                if qs.exists():
                    grouped[model._meta.verbose_name_plural.title()] = list(qs)

        # --- Create workbook ---
        wb = Workbook()
        wb.remove(wb.active)

        for model_name, objs in grouped.items():
            if not objs:
                continue
            # use your helper instead of hard truncation
            sheet_title = shorten_name(model_name)
            ws = wb.create_sheet(title=sheet_title)

            fields = objs[0]._meta.fields
            headers = [f.verbose_name.title() for f in fields]
            ws.append(headers)

            # Style headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Rows
            for obj in objs:
                row = []
                for f in fields:
                    val = getattr(obj, f.name, "")
                    if isinstance(val, datetime.datetime):
                        val = val.date()
                    if val is None:
                        val = ""
                    row.append(to_title(str(val)))
                ws.append(row)

            # Auto-size columns
            for col_num, col_cells in enumerate(ws.columns, 1):
                max_length = 0
                column = get_column_letter(col_num)
                for cell in col_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            # Freeze header
            ws.freeze_panes = "A2"

        # --- Filename with date range ---
        dates = []
        for objs in grouped.values():
            for doc in objs:
                d = getattr(doc, "date_filed", None) or getattr(doc, "date", None)
                if d:
                    if isinstance(d, datetime.datetime):
                        d = d.date()
                    dates.append(d)
        min_date, max_date = (min(dates), max(dates)) if dates else (None, None)
        date_part = f"{min_date}_{max_date}" if min_date and max_date else ""

        filename = f"Documents_{date_part}.xlsx" if date_part else "Documents.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        safe_filename = quote(filename)
        response["Content-Disposition"] = f"attachment; filename={safe_filename}; filename*=UTF-8''{safe_filename}"
        wb.save(response)
        return response
    
class UploadExcelView(View):
    """
    Initial upload endpoint - stores files in cache and returns session ID
    """
    
    @staticmethod
    def normalize_header(value):
        """Normalize headers to simplify matching."""
        return (
            str(value).strip()
            .lower()
            .replace(" ", "_")
            .replace("(", "")
            .replace(")", "")
            .replace("/", "_")
            .replace("-", "_")
        )

    def post(self, request, *args, **kwargs):
        files = request.FILES.getlist("files")

        if not files:
            return JsonResponse({
                'status': 'error',
                'message': 'No files were uploaded.'
            })

        # Get or create session ID
        session_id = request.POST.get('session_id')
        if not session_id:
            session_id = str(uuid.uuid4())

        # Store files temporarily in cache (convert to byte strings)
        file_data_list = []
        for file in files:
            file_data_list.append({
                'name': file.name,
                'content': file.read(),
            })
        
        cache.set(f'upload_files_{session_id}', file_data_list, timeout=600)

        # Initialize progress
        cache.set(f'upload_progress_{session_id}', {
            'status': 'queued',
            'current_row': 0,
            'total_rows': 0,
            'current_file': 0,
            'total_files': len(files),
            'percentage': 0,
            'message': 'Upload queued...',
            'timestamp': time.time()
        }, timeout=600)

        # Return session ID immediately
        return JsonResponse({
            'status': 'queued',
            'session_id': session_id
        })

class ProcessUploadView(View):
    """
    Separate view that processes the upload.
    This is called via AJAX and streams progress updates.
    """
    
    @staticmethod
    def normalize_header(value):
        """Normalize headers to simplify matching."""
        return (
            str(value).strip()
            .lower()
            .replace(" ", "_")
            .replace("(", "")
            .replace(")", "")
            .replace("/", "_")
            .replace("-", "_")
        )
    
    @staticmethod
    def extract_report_metadata(ws, header_row=None, last_data_row=None):
        """Extract report metadata from the worksheet."""
        metadata = {
            'dti_office': None,
            'report_no': None,
            'report_collection_date': None,
            'certification': None,
            'name_of_collecting_officer': None,
            'official_designation': None,
        }
        
        # Determine search boundary - only search BEFORE the data table header
        search_limit = header_row if header_row else min(10, ws.max_row + 1)
        
        # Search only header area for metadata (before data table)
        for row_idx in range(1, search_limit):
            for cell in ws[row_idx]:
                if not cell.value:
                    continue
                    
                cell_text = str(cell.value).strip()
                cell_lower = cell_text.lower()
                
                # Skip if this looks like the data table header row
                if 'official receipt' in cell_lower and 'number' in cell_lower:
                    continue
                
                # Extract DTI Office
                if 'dti' in cell_lower and 'office' in cell_lower:
                    metadata['dti_office'] = cell_text
                
                # Extract Report No (but NOT Official Receipt No)
                if ('report no' in cell_lower or 'report_no' in cell_lower) and 'receipt' not in cell_lower and 'official' not in cell_lower:
                    # Check if number is in same cell
                    parts = cell_text.split('.')
                    if len(parts) > 1:
                        metadata['report_no'] = parts[-1].strip()
                    else:
                        # Check adjacent cells
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            metadata['report_no'] = str(next_cell.value).strip()
                
                # Extract Date (usually below DTI Office)
                if not metadata['report_collection_date']:
                    # Check if this looks like a date
                    if isinstance(cell.value, datetime.datetime):
                        metadata['report_collection_date'] = cell.value.date()
                    elif isinstance(cell.value, str):
                        # Try to parse date formats
                        for fmt in ['%m/%d/%Y', '%m-%d-%Y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                            try:
                                parsed_date = datetime.datetime.strptime(cell_text, fmt).date()
                                metadata['report_collection_date'] = parsed_date
                                break
                            except:
                                continue
                    elif isinstance(cell.value, float) or isinstance(cell.value, int):
                        # Could be Excel date serial number
                        try:
                            if 30000 < cell.value < 60000:  # Valid Excel date range
                                base_date = datetime.datetime(1899, 12, 30)
                                parsed_date = base_date + datetime.timedelta(days=cell.value)
                                metadata['report_collection_date'] = parsed_date.date()
                        except:
                            pass
        
        # Search for certification text (in footer area AFTER data rows)
        footer_start = last_data_row if last_data_row else max(0, ws.max_row - 30)
        certification_start = None
        
        for row_idx in range(footer_start, ws.max_row + 1):
            for cell in ws[row_idx]:
                if cell.value and isinstance(cell.value, str):
                    cell_lower = str(cell.value).lower()
                    if 'i hereby certify' in cell_lower or 'certification' in cell_lower:
                        certification_start = row_idx
                        break
            if certification_start:
                break
        
        if certification_start:
            # Collect certification text (next 10 rows)
            cert_lines = []
            for row_idx in range(certification_start, min(certification_start + 10, ws.max_row + 1)):
                row_text = []
                for cell in ws[row_idx]:
                    if cell.value:
                        row_text.append(str(cell.value).strip())
                if row_text:
                    cert_lines.append(' '.join(row_text))
            
            if cert_lines:
                metadata['certification'] = ' '.join(cert_lines)
        
        # Search for collecting officer name and designation (in footer area AFTER data rows)
        for row_idx in range(footer_start, ws.max_row + 1):
            for cell in ws[row_idx]:
                if cell.value and isinstance(cell.value, str):
                    cell_lower = str(cell.value).lower()
                    
                    # Look for "Name and Signature of Collecting Officer" label
                    if 'name' in cell_lower and 'signature' in cell_lower and 'collecting officer' in cell_lower:
                        # The actual name should be in the row ABOVE (since label is below)
                        name_cell = ws.cell(row=cell.row - 1, column=cell.column)
                        if name_cell.value:
                            name_text = str(name_cell.value).strip()
                            # Skip if it looks like an underline or empty
                            if name_text and name_text not in ['_', '__', '___', '____', '_____'] and not all(c in '_ ' for c in name_text):
                                metadata['name_of_collecting_officer'] = name_text
                                print(f"DEBUG - Found officer name: {name_text}")
                        
                        # Also check 2 rows above in case there's an underline row
                        if not metadata['name_of_collecting_officer']:
                            name_cell_2 = ws.cell(row=cell.row - 2, column=cell.column)
                            if name_cell_2.value:
                                name_text = str(name_cell_2.value).strip()
                                if name_text and not all(c in '_ ' for c in name_text):
                                    metadata['name_of_collecting_officer'] = name_text
                                    print(f"DEBUG - Found officer name (2 rows up): {name_text}")
                    
                    # Look for "Official Designation" label
                    if 'official designation' in cell_lower or (cell_lower == 'official designation'):
                        # The actual designation should be in the row ABOVE
                        designation_cell = ws.cell(row=cell.row - 1, column=cell.column)
                        if designation_cell.value:
                            designation_text = str(designation_cell.value).strip()
                            # Skip if it looks like an underline or empty
                            if designation_text and designation_text not in ['_', '__', '___', '____', '_____'] and not all(c in '_ ' for c in designation_text):
                                metadata['official_designation'] = designation_text
                                print(f"DEBUG - Found designation: {designation_text}")
                        
                        # Also check 2 rows above in case there's an underline row
                        if not metadata['official_designation']:
                            designation_cell_2 = ws.cell(row=cell.row - 2, column=cell.column)
                            if designation_cell_2.value:
                                designation_text = str(designation_cell_2.value).strip()
                                if designation_text and not all(c in '_ ' for c in designation_text):
                                    metadata['official_designation'] = designation_text
                                    print(f"DEBUG - Found designation (2 rows up): {designation_text}")
                    
                    # Look for "Special Collecting Officer" as a standalone value
                    if 'special collecting officer' in cell_lower and not metadata['official_designation']:
                        metadata['official_designation'] = 'Special Collecting Officer'
                        print(f"DEBUG - Found designation: Special Collecting Officer")
        
        return metadata
    
    def get(self, request, session_id):   
        print(f"ProcessUploadView GET called with session_id: {session_id}")  # DEBUG     
        def process_and_stream():
            try:
                # Get files from cache
                file_data_list = cache.get(f'upload_files_{session_id}')
                if not file_data_list:
                    yield self._sse_message({'status': 'error', 'message': 'Files not found'})
                    return

                yield self._sse_message({
                    'status': 'processing',
                    'message': 'Starting processing...',
                    'percentage': 0
                })

                created_reports = []
                
                # First pass: Count total rows
                total_rows = 0
                file_row_counts = []
                
                yield self._sse_message({
                    'status': 'counting',
                    'message': 'Counting rows...',
                    'percentage': 0
                })
                
                for file_data in file_data_list:
                    ext = os.path.splitext(file_data['name'])[1].lower()
                    if file_data['name'].startswith("~$") or ext not in [".xlsx", ".xls"]:
                        file_row_counts.append(0)
                        continue
                    
                    try:
                        file_obj = BytesIO(file_data['content'])
                        wb = load_workbook(file_obj, data_only=True)
                        ws = wb.active
                        row_count = max(0, ws.max_row - 2)
                        file_row_counts.append(row_count)
                        total_rows += row_count
                        wb.close()
                    except Exception as e:
                        print(f"Error counting rows: {e}")
                        file_row_counts.append(0)

                yield self._sse_message({
                    'status': 'processing',
                    'current_row': 0,
                    'total_rows': total_rows,
                    'current_file': 0,
                    'total_files': len(file_data_list),
                    'percentage': 0,
                    'message': f'Processing {total_rows} rows...'
                })

                current_row = 0
                
                # Process each file
                for file_index, file_data in enumerate(file_data_list):
                    ext = os.path.splitext(file_data['name'])[1].lower()
                    
                    yield self._sse_message({
                        'status': 'processing',
                        'current_row': current_row,
                        'total_rows': total_rows,
                        'current_file': file_index + 1,
                        'total_files': len(file_data_list),
                        'percentage': round((current_row / total_rows * 100) if total_rows > 0 else 0, 2),
                        'message': f'Processing {file_data["name"]}...',
                        'current_filename': file_data['name']
                    })
                    
                    if file_data['name'].startswith("~$") or ext not in [".xlsx", ".xls"]:
                        continue

                    try:
                        file_obj = BytesIO(file_data['content'])
                        wb = load_workbook(file_obj, data_only=True)
                        ws = wb.active
                    except Exception as e:
                        print(f"Error opening file: {e}")
                        continue

                    # Auto-detect header row
                    header_row = None
                    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        cells = [str(c).strip().lower() for c in row if c]
                        if any("official receipt" in c or "date" == c for c in cells):
                            header_row = i
                            break

                    if not header_row:
                        wb.close()
                        continue

                    # Process all data rows first to find where they end
                    last_data_row = header_row + 2
                    for row_index, row in enumerate(ws.iter_rows(min_row=header_row + 2, values_only=True), start=header_row + 2):
                        if any(row):  # If row has any data
                            last_data_row = row_index
                    
                    # Extract metadata AFTER we know where the header row and last data row are
                    metadata = self.extract_report_metadata(ws, header_row, last_data_row)

                    # Combine header rows
                    headers = []
                    for c1, c2 in zip(ws[header_row], ws[header_row + 1]):
                        part1 = str(c1.value).strip() if c1.value else ""
                        part2 = str(c2.value).strip() if c2.value else ""
                        combined = f"{part1} {part2}".strip()
                        headers.append(self.normalize_header(combined))

                    # Field mapping
                    field_map = {}
                    for f in CollectionReportItem._meta.fields:
                        opts = {self.normalize_header(f.verbose_name or ""), self.normalize_header(f.name)}
                        for h in headers:
                            if h in opts:
                                field_map[h] = f.name

                    # Manual mappings - enhanced for better detection
                    if "official_receipt_number" in headers:
                        field_map["official_receipt_number"] = "number"
                    if "official_receipt_number" not in field_map and "number" in headers:
                        field_map["number"] = "number"
                    # Check for any header containing both "receipt" and "number"
                    if "number" not in field_map:
                        for h in headers:
                            if "receipt" in h and "number" in h:
                                field_map[h] = "number"
                                break
                    # Last resort: map any column called just "number"
                    if "number" not in field_map:
                        for idx, h in enumerate(headers):
                            if h == "number":
                                field_map[h] = "number"
                                break
                    
                    if "date" not in field_map and any("date" in h for h in headers):
                        date_header = next(h for h in headers if "date" in h)
                        field_map[date_header] = "date"
                    
                    print(f"DEBUG - Headers: {headers}")
                    print(f"DEBUG - Field map: {field_map}")

                    # Create report with metadata
                    report = CollectionReport.objects.create(
                        dti_office=metadata['dti_office'],
                        report_no=metadata['report_no'],
                        report_collection_date=metadata['report_collection_date'],
                        certification=metadata['certification'],
                        name_of_collecting_officer=metadata['name_of_collecting_officer'],
                        official_designation=metadata['official_designation']
                    )
                    
                    db_column_indices = [idx for idx, h in enumerate(headers) if h in field_map]
                    empty_row_count = 0
                    MAX_EMPTY_ROWS = 5
                    last_data_row = header_row + 2  # Track the last row with data

                    # Process rows
                    for row_index, row in enumerate(ws.iter_rows(min_row=header_row + 2, values_only=True), start=header_row + 2):
                        if not any(row[idx] if idx < len(row) else None for idx in db_column_indices):
                            empty_row_count += 1
                            if empty_row_count >= MAX_EMPTY_ROWS:
                                break
                            continue

                        empty_row_count = 0
                        last_data_row = row_index  # Update last row with actual data
                        current_row += 1
                        
                        # Send update every 5 rows for better performance
                        if current_row % 5 == 0:
                            percentage = round((current_row / total_rows * 100) if total_rows > 0 else 0, 2)
                            yield self._sse_message({
                                'status': 'processing',
                                'current_row': current_row,
                                'total_rows': total_rows,
                                'current_file': file_index + 1,
                                'total_files': len(file_data_list),
                                'percentage': percentage,
                                'message': f'Processing row {current_row} of {total_rows}',
                                'current_filename': file_data['name']
                            })

                        data = {}
                        for idx, value in enumerate(row):
                            header = headers[idx] if idx < len(headers) else None
                            field_name = field_map.get(header)
                            if not field_name:
                                continue

                            # Handle Date column
                            if field_name == "date":
                                if isinstance(value, datetime.datetime):
                                    value = value.date()
                                elif isinstance(value, float):
                                    base_date = datetime.datetime(1899, 12, 30)
                                    value = base_date + datetime.timedelta(days=value)
                                    value = value.date()
                                elif isinstance(value, str):
                                    try:
                                        value = datetime.datetime.strptime(value.strip(), "%Y-%m-%d").date()
                                    except:
                                        try:
                                            value = datetime.datetime.strptime(value.strip(), "%m/%d/%Y").date()
                                        except:
                                            value = None

                            data[field_name] = value
                        
                        # Extra fallback: If 'number' field is missing, try to detect it
                        if 'number' not in data or not data['number']:
                            # First, find where the date column is
                            date_col_idx = None
                            for idx, h in enumerate(headers):
                                if field_map.get(h) == 'date':
                                    date_col_idx = idx
                                    break
                            
                            # Try to find receipt number in any column
                            for idx, value in enumerate(row):
                                if not value:
                                    continue
                                
                                # Skip if this is the date column itself
                                if date_col_idx is not None and idx == date_col_idx:
                                    continue
                                    
                                value_str = str(value).strip()
                                
                                # Skip if it's clearly not a receipt number
                                if not value_str or value_str.lower() in ['none', 'n/a', '']:
                                    continue
                                
                                # Priority 1: Starts with 'BN-' or similar prefix (HIGHEST PRIORITY)
                                if value_str.startswith(('BN-', 'RN-', 'OR-', 'AR-')):
                                    data['number'] = value_str
                                    print(f"DEBUG - Found receipt number (prefix): {value_str}")
                                    break
                                
                                # Priority 2: Contains multiple hyphens (like BN-25-0140-0061-0009411-0009444)
                                if value_str.count('-') >= 2 and len(value_str) > 10:
                                    data['number'] = value_str
                                    print(f"DEBUG - Found receipt number (hyphens): {value_str}")
                                    break
                            
                            # Priority 3: If still not found, look for number to the right of date column
                            if 'number' not in data or not data['number']:
                                if date_col_idx is not None:
                                    # Check column immediately to the right of date
                                    for offset in [1, 2]:  # Check next 2 columns
                                        check_idx = date_col_idx + offset
                                        if check_idx < len(row):
                                            value = row[check_idx]
                                            if not value:
                                                continue
                                            
                                            value_str = str(value).strip()
                                            
                                            # Is it a long whole number?
                                            try:
                                                if isinstance(value, (int, float)):
                                                    # Excel dates are typically 40000-50000 range, receipt numbers are much larger
                                                    if value == int(value) and int(value) > 100000:
                                                        data['number'] = str(int(value))
                                                        print(f"DEBUG - Found receipt number (long int right of date): {value}")
                                                        break
                                                elif value_str.isdigit() and len(value_str) >= 6:
                                                    data['number'] = value_str
                                                    print(f"DEBUG - Found receipt number (digit string right of date): {value_str}")
                                                    break
                                            except:
                                                continue

                        if not data:
                            continue

                        try:
                            with transaction.atomic():
                                item = CollectionReportItem.objects.create(**data)
                                report.report_items.add(item)
                        except Exception as e:
                            print(f"Failed to save row {row_index}: {e}")
                            continue

                    created_reports.append(report)
                    wb.close()

                # Send final completion message
                redirect_url = reverse("collection-report", args=[created_reports[-1].pk]) if created_reports else reverse("all-documents")
                
                yield self._sse_message({
                    'status': 'complete',
                    'current_row': total_rows,
                    'total_rows': total_rows,
                    'current_file': len(file_data_list),
                    'total_files': len(file_data_list),
                    'percentage': 100,
                    'message': 'Upload complete!',
                    'redirect_url': redirect_url
                })

                # Clean up cache
                cache.delete(f'upload_files_{session_id}')

            except Exception as e:
                print(f"Processing error: {e}")
                import traceback
                traceback.print_exc()
                yield self._sse_message({
                    'status': 'error',
                    'message': f'Error: {str(e)}'
                })
        
        response = StreamingHttpResponse(
            process_and_stream(),
            content_type='text/event-stream'
        )
        response['Cache-Control'] = 'no-cache'
        response['X-Accel-Buffering'] = 'no'
        return response
    
    def _sse_message(self, data):
        """Format data as SSE message"""
        return f"data: {json.dumps(data)}\n\n"

class UploadProgressStreamView(View):
    """Legacy view - kept for backwards compatibility but not used in new flow"""
    
    def get(self, request, session_id):
        progress_data = cache.get(f'upload_progress_{session_id}')
        return JsonResponse(progress_data or {'status': 'not_found'})
    
class CancelUploadView(View):
    """Cancel an ongoing upload"""
    
    def post(self, request, session_id):
        # Set cancellation flag
        cache.set(f'upload_cancel_{session_id}', True, timeout=600)
        return JsonResponse({'status': 'cancellation_requested'})