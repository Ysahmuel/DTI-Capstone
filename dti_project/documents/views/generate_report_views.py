import datetime
import re
from urllib.parse import quote
from django.views import View
from django.apps import apps
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from zipfile import ZipFile
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, Image, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
import matplotlib
matplotlib.use('Agg')  # Non-GUI backend suitable for server-side plotting
import matplotlib.pyplot as plt
from reportlab.platypus import Image, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import PageBreak
from reportlab.lib.pagesizes import landscape, A4
from matplotlib.ticker import MaxNLocator
from ..mixins.filter_mixins import FilterableDocumentMixin
from ..models.base_models import DraftModel
from ..mixins.permissions_mixins import UserRoleMixin
from ..models import (
    ChecklistEvaluationSheet,
    InspectionValidationReport,
    OrderOfPayment,
    PersonalDataSheet,
    SalesPromotionPermitApplication,
    ServiceRepairAccreditationApplication,
)
    
MODEL_MAP = {
    model._meta.model_name: model
    for model in apps.get_models()
    if issubclass(model, DraftModel) and not model._meta.abstract
}

EXPORT_MODEL_MAP = {
    "salespromotionpermitapplication": SalesPromotionPermitApplication,
    "personaldatasheet": PersonalDataSheet,
    "servicerepairaccreditationapplication": ServiceRepairAccreditationApplication,
    "inspectionvalidationreport": InspectionValidationReport,
    "orderofpayment": OrderOfPayment,
    "checklistevaluationsheet": ChecklistEvaluationSheet,
}

def to_title(value):
    """Normalize strings: remove underscores, title-case, handle non-strings gracefully."""
    if not isinstance(value, str):
        return value
    return value.replace("_", " ").title()

def clean_sheet_name(name: str) -> str:
    """
    Format sheet name with spaces, title case, and Excel-safe rules.
    """
    # Normalize underscores and multiple spaces
    name = re.sub(r'[_\s]+', ' ', name)

    # Remove forbidden Excel characters
    name = re.sub(r'[:\\/?*\[\]]', '', name)

    # Title case
    name = to_title(name.strip())

    # Excel max length = 31 chars
    return name[:31]

class GenerateDocumentsReportView(LoginRequiredMixin, FilterableDocumentMixin, View):

    def export_excel(self):
        from openpyxl.utils import get_column_letter

        MODEL_MAP = {
            model._meta.model_name: model
            for model in apps.get_models()
            if issubclass(model, DraftModel) and not model._meta.abstract
        }

        wb = Workbook()
        wb.remove(wb.active)
        included_models = []

        for model_name, model in MODEL_MAP.items():
            qs = model.objects.all()
            filtered_qs = self.apply_filters(qs)
            objs = list(filtered_qs)
            if not objs:
                continue

            model_verbose_name = model._meta.verbose_name
            included_models.append(model_verbose_name)

            sheet_title = clean_sheet_name(model_verbose_name)
            ws = wb.create_sheet(title=sheet_title)

            fields = [f for f in model._meta.get_fields() if f.concrete and not f.many_to_many and f.name != 'id']
            headers = [to_title(f.verbose_name) for f in fields]
            num_cols = len(headers)

            # Header section
            ws.merge_cells('F1:I1'); ws['F1'].value = f"{to_title(model_verbose_name)} Report"; ws['F1'].font = Font(bold=True, size=12)
            ws.merge_cells('F2:I2'); ws['F2'].value = "DTI Albay Provincial Office"; ws['F2'].font = Font(bold=True, size=10)
            ws.merge_cells('F3:I3'); ws['F3'].value = datetime.date.today().strftime("%B %d, %Y"); ws['F3'].font = Font(size=10)
            ws.row_dimensions[4].height = 10
            thick_border = Border(bottom=Side(style='thick', color='000000'))
            for col_num in range(1, num_cols + 1):
                ws.cell(row=4, column=col_num).border = thick_border

            # Column headers row 5
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(top=Side(style='thin', color='000000'),
                                     bottom=Side(style='thin', color='000000'),
                                     left=Side(style='thin', color='000000'),
                                     right=Side(style='thin', color='000000'))

            # Data rows
            current_row = 6
            for obj in objs:
                row = []
                for f in fields:
                    val = getattr(obj, f.name, "")
                    if isinstance(val, datetime.datetime):
                        val = val.date()
                    if val is None:
                        val = ""
                    row.append(to_title(str(val)) if val != "" else "")
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = value
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = Border(top=Side(style='thin', color='CCCCCC'),
                                         bottom=Side(style='thin', color='CCCCCC'),
                                         left=Side(style='thin', color='CCCCCC'),
                                         right=Side(style='thin', color='CCCCCC'))
                current_row += 1

            # Auto-fit columns
            for col_num in range(1, num_cols + 1):
                max_length = 0
                column_letter = get_column_letter(col_num)
                for row_num in range(5, current_row):
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)

        # Fallback sheet
        if not wb.sheetnames:
            ws = wb.create_sheet("No Data")
            ws['A1'] = "No Data Available"
            ws['A1'].font = Font(size=10)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer

    def export_pdf(self):
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30
        )
        elements = []
        styles = getSampleStyleSheet()

        # Header
        elements.append(Paragraph("Documents Report", styles['Title']))
        elements.append(Paragraph("DTI Albay Provincial Office", styles['Heading2']))
        elements.append(Paragraph(f"Generated on: {datetime.date.today().strftime('%B %d, %Y')}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Collect document info
        MODEL_MAP = {
            model._meta.model_name: model
            for model in apps.get_models()
            if issubclass(model, DraftModel) and not model._meta.abstract
        }

        doc_type_counts = {}
        status_counts = {}
        date_samples = []

        DOC_SPECIFIC_FIELDS = {
            "salespromotionpermitapplication": ["coverage"],
            "inspectionvalidationreport": ["type_of_application_activity"],
            "checklistevaluationsheet": ["type_of_application", "star_rating"],
            "servicerepairaccreditationapplication": [
                "application_type", "category", "star_rating",
                "social_classification", "asset_size", "form_of_organization"
            ],
            "personaldatasheet": ["sex", "civil_status", "nationality"],
        }

        doc_specific_data = {}
        preferred_date_fields = ["created_at", "submitted_at", "date", "date_filed", "date_of_application"]

        # --- Collect data ---
        for model_name, model in MODEL_MAP.items():
            qs = self.apply_filters(model.objects.all())
            objs = list(qs)
            if not objs:
                continue

            model_verbose_name = to_title(model._meta.verbose_name)
            doc_type_counts[model_verbose_name] = len(objs)

            # Collect one date per object
            model_date_fields = [f.name for f in model._meta.fields if f.get_internal_type() in ['DateField', 'DateTimeField']]
            for o in objs:
                chosen_date = None
                for p in preferred_date_fields:
                    if p in model_date_fields:
                        chosen_date = getattr(o, p, None)
                        if chosen_date:
                            break
                if not chosen_date and model_date_fields:
                    chosen_date = getattr(o, model_date_fields[0], None)
                if chosen_date:
                    if isinstance(chosen_date, datetime.datetime):
                        chosen_date = chosen_date.date()
                    date_samples.append(chosen_date)

            # Status counts
            if "status" in [f.name for f in model._meta.fields]:
                for o in objs:
                    status = getattr(o, "status", "Unknown")
                    status_counts[status] = status_counts.get(status, 0) + 1

            # Document-specific fields
            key = model_name.lower()
            if key in DOC_SPECIFIC_FIELDS:
                for field in DOC_SPECIFIC_FIELDS[key]:
                    values_count = {}
                    for o in objs:
                        val = getattr(o, field, None)
                        if val is not None:
                            if field == "star_rating":
                                val = f"{val} Star"
                            values_count[val] = values_count.get(val, 0) + 1
                    doc_specific_data[f"{model_verbose_name} - {to_title(field)}"] = values_count

        # --- Sample size paragraph ---
        if date_samples:
            oldest = min(date_samples)
            newest = max(date_samples)
            elements.append(Paragraph(
                f"Sample Size: {len(date_samples)} documents ({oldest} - {newest})",
                styles['Normal']
            ))
            elements.append(Spacer(1, 12))

        # --- Pie chart helper ---
        def build_pie_chart(counts_dict):
            labels = list(counts_dict.keys())
            sizes = list(counts_dict.values())
            cmap = plt.get_cmap("tab20")
            colors_list = [cmap(i) for i in range(len(sizes))]
            hex_colors = [f'#{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}' for r, g, b, _ in colors_list]

            fig, ax = plt.subplots(figsize=(2.5, 2.5))
            ax.pie(sizes, labels=None, colors=colors_list, startangle=90)
            ax.axis('equal')
            plt.tight_layout(pad=0.1)
            buf = BytesIO()
            plt.savefig(buf, format='PNG', bbox_inches='tight')
            plt.close(fig)
            buf.seek(0)

            bullet_style = ParagraphStyle('BulletStyle', fontSize=9, leftIndent=10, spaceAfter=2)
            total = sum(sizes)
            bullets = [
                Paragraph(
                    f'<font color="{hex_colors[i]}">&#9679;</font> {labels[i]}: {sizes[i]} ({sizes[i] / total * 100:.1f}%)',
                    bullet_style
                )
                for i in range(len(labels))
            ]
            return buf, bullets

        # --- Combine pie charts ---
        charts_data = []
        if doc_type_counts:
            buf, bullets = build_pie_chart(doc_type_counts)
            charts_data.append((buf, bullets, "Document Types"))
        if status_counts:
            buf, bullets = build_pie_chart(status_counts)
            charts_data.append((buf, bullets, "Statuses"))
        for title, data in doc_specific_data.items():
            if data:
                buf, bullets = build_pie_chart(data)
                charts_data.append((buf, bullets, title))

        # --- Layout: 2 charts per row ---
        for i in range(0, len(charts_data), 2):
            row_data = []
            for j in range(2):
                if i + j < len(charts_data):
                    buf, bullets, title = charts_data[i + j]
                    if " - " in title:
                        doc_title, field_title = title.split(" - ", 1)
                        main_heading = Paragraph(doc_title, styles['Heading3'])
                        small_heading_style = ParagraphStyle(
                            'SmallHeading',
                            parent=styles['Normal'],
                            fontName='Helvetica',
                            fontSize=9,
                            leading=11,
                            spaceAfter=6,
                        )
                        sub_heading = Paragraph(field_title, small_heading_style)
                        chart_table = Table([
                            [main_heading],
                            [sub_heading],
                            [Image(buf, width=150, height=150)],
                            [bullets],
                        ])
                    else:
                        chart_table = Table([
                            [Paragraph(title, styles['Heading3'])],
                            [Image(buf, width=150, height=150)],
                            [bullets],
                        ])
                    chart_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
                    row_data.append(chart_table)
                else:
                    row_data.append("")
            page_table = Table([row_data], colWidths=[doc.width / 2.0] * 2)
            page_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
            elements.append(page_table)
            elements.append(Spacer(1, 24))


        # --- Adaptive bar chart based on date range ---
        if date_samples:
            min_date, max_date = min(date_samples), max(date_samples)
            delta_days = (max_date - min_date).days + 1

            from collections import Counter

            # Daily counts
            if delta_days < 7:
                counts_dict = Counter(date_samples)
                x_labels = sorted(counts_dict.keys())
                y_values = [counts_dict[d] for d in x_labels]
                x_labels_fmt = [d.strftime("%b %d") for d in x_labels]
                chart_title = "Documents per Day"

            # Weekly counts
            elif delta_days < 28:
                weekly_counts = Counter()
                for d in date_samples:
                    year, week, _ = d.isocalendar()
                    weekly_counts[(year, week)] += 1

                sorted_weeks = sorted(weekly_counts.keys())
                x_labels_fmt = []
                y_values = []
                for year, week in sorted_weeks:
                    start_date = datetime.date.fromisocalendar(year, week, 1)
                    end_date = datetime.date.fromisocalendar(year, week, 7)
                    x_labels_fmt.append(f"{start_date.strftime('%b %d')}–{end_date.strftime('%b %d')}")
                    y_values.append(weekly_counts[(year, week)])
                chart_title = "Documents per Week"

            # Monthly counts
            else:
                monthly_counts = Counter()
                for d in date_samples:
                    monthly_counts[d.strftime("%Y-%m")] += 1
                sorted_months = sorted(monthly_counts.keys())
                x_labels_fmt = sorted_months
                y_values = [monthly_counts[m] for m in sorted_months]
                chart_title = "Documents per Month"

            # --- Plot bar chart ---
            num_bars = len(x_labels_fmt)
            fig_width = max(10, num_bars * 0.8)
            fig, ax = plt.subplots(figsize=(fig_width, 5))
            bar_width = 0.6 if num_bars > 1 else 0.3

            ax.bar(range(num_bars), y_values, color='skyblue', width=bar_width)
            ax.set_title(chart_title, fontsize=16)
            ax.set_xlabel("Time Period")
            ax.set_ylabel("Total Documents")
            ax.set_xticks(range(num_bars))
            ax.set_xticklabels(x_labels_fmt, rotation=45, ha='right')

            # Only integer y-axis
            ax.yaxis.set_major_locator(MaxNLocator(integer=True))

            plt.tight_layout(pad=2.0)
            buf = BytesIO()
            plt.savefig(buf, format='PNG')
            plt.close(fig)
            buf.seek(0)

            elements.append(Spacer(1, 24))
            elements.append(Image(buf, width=doc.width, height=doc.height*0.5))
            elements.append(Spacer(1, 24))

        # --- Build the PDF ---
        doc.build(elements) 
        buffer.seek(0)       
        return buffer        

    def post(self, request, *args, **kwargs):
        excel_buffer = self.export_excel()
        pdf_buffer = self.export_pdf()

        zip_buffer = BytesIO()
        with ZipFile(zip_buffer, 'w') as zip_file:
            zip_file.writestr(f"Documents Report {datetime.date.today()}.xlsx", excel_buffer.getvalue())
            zip_file.writestr(f"Documents Report {datetime.date.today()}.pdf", pdf_buffer.getvalue())

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="Documents Report {datetime.date.today()}.zip"'
        return response