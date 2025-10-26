import datetime
from urllib.parse import quote
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ..models.collection_models import CollectionReport, CollectionReportItem
from django.contrib.auth.mixins import LoginRequiredMixin
from ..mixins.permissions_mixins import UserRoleMixin
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ExportDocumentsExcelView(LoginRequiredMixin, View):
    """
    Export collection reports in the standardized DTI format matching the template.
    """

    def post(self, request, *args, **kwargs):
        user = request.user
        selected_docs = request.POST.getlist("documents")

        # --- Collect CollectionReports ---
        reports = []
        if selected_docs:
            # Filter for collection reports only
            report_ids = []
            for doc in selected_docs:
                try:
                    model_name, pk = doc.split(":")
                    if model_name.lower() == "collectionreport":
                        report_ids.append(pk)
                except ValueError:
                    continue
            
            if report_ids:
                qs = UserRoleMixin.get_queryset_or_all(CollectionReport, user).filter(pk__in=report_ids)
                reports = list(qs)
        else:
            # Export all collection reports
            qs = UserRoleMixin.get_queryset_or_all(CollectionReport, user)
            reports = list(qs)

        if not reports:
            # Return empty workbook if no reports
            wb = Workbook()
            ws = wb.active
            ws.title = "No Data"
            ws.append(["No collection reports found to export"])
            
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f"attachment; filename=No_Reports.xlsx"
            wb.save(response)
            return response

        # --- Create workbook ---
        wb = Workbook()
        wb.remove(wb.active)

        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for report in reports:
            # Create sheet name from report number or date
            sheet_name = f"Report {report.report_no}" if report.report_no else f"Report {report.pk}"
            sheet_name = sheet_name[:31]  # Excel sheet name limit
            ws = wb.create_sheet(title=sheet_name)

            current_row = 1

            # --- HEADER SECTION ---
            # Title
            ws.merge_cells(f'A{current_row}:H{current_row}')
            title_cell = ws[f'A{current_row}']
            title_cell.value = "REPORT OF COLLECTIONS AND DEPOSITS"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = center_alignment
            current_row += 1

            # DTI Office
            ws.merge_cells(f'A{current_row}:H{current_row}')
            office_cell = ws[f'A{current_row}']
            office_cell.value = report.dti_office or "DTI Provincial Office"
            office_cell.font = Font(bold=True)
            office_cell.alignment = center_alignment
            current_row += 1

            # Date
            ws.merge_cells(f'A{current_row}:H{current_row}')
            date_cell = ws[f'A{current_row}']
            date_cell.value = report.report_collection_date or datetime.date.today()
            date_cell.alignment = center_alignment
            current_row += 1

            # Report Number
            ws.merge_cells(f'A{current_row}:H{current_row}')
            report_no_cell = ws[f'A{current_row}']
            report_no_cell.value = f"Report No. {report.report_no}" if report.report_no else ""
            report_no_cell.alignment = center_alignment
            current_row += 1

            # Blank row
            current_row += 1

            # --- TABLE HEADER ---
            header_row = current_row
            
            # First header row (merged cells for compound headers)
            ws.merge_cells(f'A{header_row}:B{header_row}')
            ws[f'A{header_row}'] = "Official Receipt"
            ws[f'C{header_row}'] = "Responsibility"
            ws[f'D{header_row}'] = ""
            ws[f'E{header_row}'] = ""
            ws[f'F{header_row}'] = "Particulars"
            ws[f'G{header_row}'] = ""
            ws[f'H{header_row}'] = "Amount"
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                cell = ws[f'{col}{header_row}']
                cell.font = bold_font
                cell.alignment = center_alignment
                cell.fill = header_fill
                cell.border = thin_border
            
            current_row += 1

            # Second header row (subheaders)
            ws[f'A{current_row}'] = "Date"
            ws[f'B{current_row}'] = "Number"
            ws[f'C{current_row}'] = "Code"
            ws[f'D{current_row}'] = ""
            ws[f'E{current_row}'] = "Payor"
            ws[f'F{current_row}'] = ""
            ws[f'G{current_row}'] = ""
            ws[f'H{current_row}'] = ""
            
            for col in ['A', 'B', 'C', 'E', 'H']:
                cell = ws[f'{col}{current_row}']
                cell.font = bold_font
                cell.alignment = center_alignment
                cell.fill = header_fill
                cell.border = thin_border
            
            current_row += 1

            # --- DATA ROWS ---
            items = report.report_items.all().order_by('date', 'number')
            data_start_row = current_row
            
            for item in items:
                ws[f'A{current_row}'] = item.date if item.date else ""
                ws[f'B{current_row}'] = item.number if item.number else ""
                ws[f'C{current_row}'] = item.responsibility_code if item.responsibility_code else ""
                ws[f'D{current_row}'] = ""
                ws[f'E{current_row}'] = item.payor if item.payor else ""
                ws[f'F{current_row}'] = item.particulars if item.particulars else ""
                ws[f'G{current_row}'] = ""
                ws[f'H{current_row}'] = item.amount if item.amount else 0
                
                # Apply borders and alignment
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    cell = ws[f'{col}{current_row}']
                    cell.border = thin_border
                    if col in ['A', 'B', 'C']:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col == 'H':
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0.00'
                
                current_row += 1

            # --- SUMMARY SECTION ---
            current_row += 2  # Add blank rows
            
            ws.merge_cells(f'A{current_row}:D{current_row}')
            ws[f'A{current_row}'] = "Summary:"
            ws[f'A{current_row}'].font = bold_font
            current_row += 1

            # Undeposited Collections
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = "Undeposited Collections per last Report"
            ws[f'F{current_row}'] = "P"
            ws[f'H{current_row}'] = 0  # This would need to be calculated or stored
            ws[f'H{current_row}'].number_format = '#,##0.00'
            current_row += 2

            # Collections per OR
            ws.merge_cells(f'A{current_row}:E{current_row}')
            or_numbers = [item.number for item in items if item.number]
            or_range = f"Collections per OR Nos. {or_numbers[0]} to {or_numbers[-1]}" if or_numbers else "Collections per OR Nos."
            ws[f'A{current_row}'] = or_range
            current_row += 1
            
            # OR number details (you can add multiple lines if needed)
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = "and " + ", ".join(or_numbers[:5]) if len(or_numbers) > 5 else ""
            ws[f'H{current_row}'] = sum(item.amount for item in items if item.amount)
            ws[f'H{current_row}'].number_format = '#,##0.00'
            current_row += 1

            # Total
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = "Total"
            ws[f'A{current_row}'].font = bold_font
            total_amount = sum(item.amount for item in items if item.amount)
            ws[f'H{current_row}'] = total_amount
            ws[f'H{current_row}'].number_format = '#,##0.00'
            ws[f'H{current_row}'].font = bold_font
            current_row += 1

            # Deposits
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = "Deposits"
            ws[f'A{current_row}'].font = bold_font
            current_row += 1

            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = "Date"
            ws[f'H{current_row}'] = ""  # Deposit amounts would go here
            ws[f'H{current_row}'].number_format = '#,##0.00'
            current_row += 3

            # Undeposited Collections this Report
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = "Undeposited Collections, this Report"
            ws[f'F{current_row}'] = "P"
            ws[f'H{current_row}'] = total_amount  # Simplified - would need actual calculation
            ws[f'H{current_row}'].number_format = '#,##0.00'
            ws[f'H{current_row}'].font = bold_font
            current_row += 2

            # --- CERTIFICATION SECTION ---
            ws.merge_cells(f'A{current_row}:H{current_row}')
            cert_cell = ws[f'A{current_row}']
            cert_cell.value = "CERTIFICATION"
            cert_cell.font = Font(bold=True, size=11)
            cert_cell.alignment = center_alignment
            current_row += 2

            # Certification text
            certification_text = report.certification if report.certification else \
                "I hereby certify that the above is a true and correct report of collections made by me during the period covered."
            
            ws.merge_cells(f'A{current_row}:H{current_row + 2}')
            cert_text_cell = ws[f'A{current_row}']
            cert_text_cell.value = certification_text
            cert_text_cell.alignment = Alignment(wrap_text=True, vertical="top")
            current_row += 4

            # Officer name
            ws.merge_cells(f'A{current_row}:D{current_row}')
            name_cell = ws[f'A{current_row}']
            name_cell.value = report.name_of_collection_officer or ""
            name_cell.font = bold_font
            name_cell.alignment = center_alignment
            current_row += 1

            ws.merge_cells(f'A{current_row}:D{current_row}')
            ws[f'A{current_row}'] = "Name and Signature of Collecting Officer"
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 2

            # Official designation
            ws.merge_cells(f'A{current_row}:D{current_row}')
            designation_cell = ws[f'A{current_row}']
            designation_cell.value = report.official_designation or "Special Collecting Officer"
            designation_cell.font = bold_font
            designation_cell.alignment = center_alignment
            current_row += 1

            ws.merge_cells(f'A{current_row}:D{current_row}')
            ws[f'A{current_row}'] = "Official Designation"
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1

            # Date
            ws.merge_cells(f'A{current_row}:D{current_row}')
            date_cert_cell = ws[f'A{current_row}']
            date_cert_cell.value = report.report_collection_date or datetime.date.today()
            date_cert_cell.alignment = center_alignment

            # --- SET COLUMN WIDTHS ---
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 5
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 35
            ws.column_dimensions['G'].width = 5
            ws.column_dimensions['H'].width = 15

        # --- GENERATE FILENAME ---
        dates = []
        for report in reports:
            d = report.report_collection_date
            if d:
                if isinstance(d, datetime.datetime):
                    d = d.date()
                dates.append(d)
        
        min_date, max_date = (min(dates), max(dates)) if dates else (None, None)
        date_part = f"{min_date}_{max_date}" if min_date and max_date else ""
        filename = f"Collection_Report_{date_part}.xlsx" if date_part else "Collection_Report.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        safe_filename = quote(filename)
        response["Content-Disposition"] = f"attachment; filename={safe_filename}; filename*=UTF-8''{safe_filename}"
        wb.save(response)
        return response